from __future__ import annotations

import json
import os
import threading
import time
from io import BytesIO
from zipfile import ZipFile

os.environ["CLASSASSIGN_USERNAME"] = "test-admin"
os.environ["CLASSASSIGN_PASSWORD"] = "test-password-only"
os.environ["CLASSASSIGN_SECRET"] = "test-session-secret-that-is-longer-than-32-characters"

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from app.assignment_rules import apply_required_teacher_assignments
from app import main
from app.store import StateStore, StorageError, UpstashStateStore


def _allocation_teacher_ids(value: list[dict]) -> list[str]:
    return [item["teacher_id"] for item in value]


def _first_teacher_id(value: list[dict]) -> str | None:
    return value[0]["teacher_id"] if value else None


def _login(client: TestClient) -> None:
    response = client.post("/api/auth/login", json={"username": "test-admin", "password": "test-password-only"})
    assert response.status_code == 200


def _generate_and_wait(client: TestClient, payload: dict, timeout: float = 30.0) -> dict:
    started = client.post("/api/schedule/generate", json=payload)
    assert started.status_code == 202, started.text
    job_id = started.json()["job"]["id"]
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        status_response = client.get("/api/schedule/generation")
        assert status_response.status_code == 200, status_response.text
        data = status_response.json()
        assert data["job"]["id"] == job_id
        if data["job"]["status"] in {"completed", "error"}:
            return data
        time.sleep(0.01)
    raise AssertionError(f"排课任务 {job_id} 未在 {timeout} 秒内完成")


def _teacher_import_file(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "教师名单"
    sheet.append(["教师姓名", "期望最低周课时", "是否班主任", "班主任班级", "任课配置"])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _configure_required_teachers(state: dict) -> None:
    for class_item in state["classes"]:
        teacher_id = f"t_required_{class_item['id']}"
        state["teachers"].append({
            "id": teacher_id,
            "name": f"{class_item['name']}班主任",
            "subject_ids": ["chinese"],
            "min_weekly_lessons": 0,
            "homeroom_class_id": class_item["id"],
        })
        state["assignments"][class_item["id"]]["chinese"] = teacher_id
    apply_required_teacher_assignments(state)


def _page_contains_image(page) -> bool:
    resources = page.get("/Resources")
    if not resources:
        return False
    xobjects = resources.get_object().get("/XObject")
    if not xobjects:
        return False
    return any(item.get_object().get("/Subtype") == "/Image" for item in xobjects.get_object().values())


def test_login_state_generate_and_export(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    configured_state = local_store.load()
    _configure_required_teachers(configured_state)
    local_store.save(configured_state)
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)

    assert client.get("/health").json() == {"status": "ok"}
    assert client.get("/api/state").status_code == 401

    bad = client.post("/api/auth/login", json={"username": "test-admin", "password": "wrong"})
    assert bad.status_code == 401
    login = client.post("/api/auth/login", json={"username": "test-admin", "password": "test-password-only"})
    assert login.status_code == 200

    state = client.get("/api/state")
    assert state.status_code == 200
    assert len(state.json()["state"]["classes"]) == 36
    assert set(state.json()["state"]["class_counts"].values()) == {6}

    generated = _generate_and_wait(client, {"seed": 202508, "attempts": 12})
    assert generated["job"]["status"] == "completed"
    assert generated["job"]["result"]["success"]

    exported = client.get("/api/export/schedule.pdf")
    assert exported.status_code == 200
    assert exported.content.startswith(b"%PDF")
    assert len(exported.content) < 2_000_000
    assert exported.headers["content-type"] == "application/pdf"
    full_pdf = PdfReader(BytesIO(exported.content))
    assert len(full_pdf.pages) == 36
    first_page_text = full_pdf.pages[0].extract_text() or ""
    assert "高唐县民族实验小学一年级1班课程表" in first_page_text
    assert "一年级1班班主任" in first_page_text
    assert float(full_pdf.pages[0].mediabox.height) > float(full_pdf.pages[0].mediabox.width)
    assert _page_contains_image(full_pdf.pages[0])

    grade_export = client.get("/api/export/schedule.pdf?grade=3")
    assert grade_export.status_code == 200
    assert len(PdfReader(BytesIO(grade_export.content)).pages) == 6
    assert client.get("/api/export/schedule.pdf?grade=7").status_code == 422

    grade_xlsx = client.get("/api/export/schedule.xlsx?grade=3")
    assert grade_xlsx.status_code == 200
    assert grade_xlsx.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    class_workbook = load_workbook(BytesIO(grade_xlsx.content))
    assert len(class_workbook.sheetnames) == 6
    first_class_sheet = class_workbook[class_workbook.sheetnames[0]]
    assert first_class_sheet["A1"].value == "高唐县民族实验小学三年级1班课程表"
    assert first_class_sheet.page_setup.orientation == "portrait"
    assert first_class_sheet.page_setup.paperSize == 9
    assert first_class_sheet.page_setup.fitToWidth == 1
    assert first_class_sheet.page_setup.fitToHeight == 1
    assert first_class_sheet.print_area == "'三年级1班'!$A$1:$G$9"
    assert first_class_sheet.row_dimensions[3].height == 92
    assert first_class_sheet.column_dimensions["C"].width == 16.5
    assert "A1:G1" in {str(item) for item in first_class_sheet.merged_cells.ranges}
    assert all(cell.fill.fill_type is None for row in first_class_sheet.iter_rows() for cell in row)


def test_schedule_generation_runs_in_background_and_deduplicates_clicks(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    monkeypatch.setattr(main, "store", local_store)
    worker_started = threading.Event()
    release_worker = threading.Event()

    def slow_generation(_state: dict, seed: int | None = None, attempts: int = 80) -> dict:
        worker_started.set()
        release_worker.wait(5)
        return {"success": False, "seed": seed, "attempts": attempts, "errors": ["测试结束"], "warnings": []}

    monkeypatch.setattr(main, "generate_schedule", slow_generation)
    client = TestClient(main.app)
    _login(client)
    try:
        first = client.post("/api/schedule/generate", json={"seed": 7, "attempts": 2})
        assert first.status_code == 202
        assert worker_started.wait(2)
        first_job = first.json()["job"]

        duplicate = client.post("/api/schedule/generate", json={"seed": 8, "attempts": 3})
        assert duplicate.status_code == 202
        assert duplicate.json()["already_running"] is True
        assert duplicate.json()["job"]["id"] == first_job["id"]
        assert client.get("/api/schedule/generation").json()["job"]["status"] == "running"
    finally:
        release_worker.set()

    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        finished = client.get("/api/schedule/generation").json()
        if finished["job"]["status"] == "completed":
            break
        time.sleep(0.01)
    else:
        raise AssertionError("后台排课测试任务未完成")
    assert finished["job"]["result"]["seed"] == 7


def test_frontend_assets_are_versioned_and_not_cached(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)

    index_response = client.get("/")
    assert index_response.status_code == 200
    assert "/assets/styles.css?v=20260827-2" in index_response.text
    assert "/assets/app.js?v=20260831-1" in index_response.text
    assert '<span class="login-school-name">高唐县民族实验小学</span>' in index_response.text
    assert 'value="teacher">指定教师' in index_response.text
    assert "teacher-export-class" not in index_response.text
    assert "批量导出课表" in index_response.text
    assert "导出教师课表" in index_response.text
    assert 'name="class-export-format"' in index_response.text
    assert 'name="teacher-export-format"' in index_response.text
    assert 'id="schedule-edit-button"' in index_response.text
    assert 'data-view="assignments"' not in index_response.text
    assert "教师与任课" in index_response.text
    assert '<img class="print-logo" src="/static/logo.jpg"' in index_response.text
    assert 'id="add-teacher-range"' in index_response.text
    assert 'id="teacher-range-list"' in index_response.text
    assert 'id="teacher-select-page"' in index_response.text
    assert 'id="teacher-bulk-delete"' in index_response.text
    assert 'id="teacher-pagination"' in index_response.text
    assert "主授" not in index_response.text
    assert "no-store" in index_response.headers["cache-control"]

    script_response = client.get("/assets/app.js?v=20260831-1")
    assert script_response.status_code == 200
    assert "no-store" in script_response.headers["cache-control"]
    assert "`${appData.state.school_name}教师课程表`" in script_response.text
    assert "teaching_assignments: teachingAssignments" in script_response.text
    assert "function teacherRangesFromState" in script_response.text
    assert "const TEACHERS_PER_PAGE = 10" in script_response.text
    assert 'api("/api/teachers/bulk-delete"' in script_response.text
    assert 'api("/api/schedule/generation")' in script_response.text
    assert "正在后台编排课表" in script_response.text
    assert "主授" not in script_response.text
    assert "data-assignment-primary" not in script_response.text

    style_response = client.get("/assets/styles.css?v=20260827-2")
    assert style_response.status_code == 200
    assert ".login-school-name { white-space: nowrap; }" in style_response.text
    assert ".teacher-range-row" in style_response.text
    assert "no-store" in style_response.headers["cache-control"]


def test_single_teacher_can_save_assignment_ranges_atomically(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)

    created = client.post(
        "/api/teachers",
        json={
            "name": "范围任课教师",
            "subject_ids": [],
            "min_weekly_lessons": 0,
            "homeroom_class_id": None,
            "teaching_assignments": {
                "g1c1": ["chinese", "math"],
                "g1c2": ["chinese"],
            },
        },
    )
    assert created.status_code == 200, created.text
    teacher = created.json()["state"]["teachers"][0]
    teacher_id = teacher["id"]
    assert set(teacher["subject_ids"]) == {"chinese", "math"}
    created_assignments = created.json()["state"]["assignments"]
    assert _first_teacher_id(created_assignments["g1c1"]["chinese"]) == teacher_id
    assert _first_teacher_id(created_assignments["g1c1"]["math"]) == teacher_id
    assert _first_teacher_id(created_assignments["g1c1"]["reading"]) == teacher_id
    assert _first_teacher_id(created_assignments["g1c2"]["chinese"]) == teacher_id
    assert _first_teacher_id(created_assignments["g1c2"]["reading"]) == teacher_id

    updated = client.put(
        f"/api/teachers/{teacher_id}",
        json={
            "name": "范围任课教师",
            "subject_ids": [],
            "min_weekly_lessons": 0,
            "homeroom_class_id": None,
            "teaching_assignments": {
                "g1c2": ["math"],
                "g1c3": ["pe"],
            },
        },
    )
    assert updated.status_code == 200, updated.text
    updated_state = updated.json()["state"]
    updated_teacher = next(item for item in updated_state["teachers"] if item["id"] == teacher_id)
    assert set(updated_teacher["subject_ids"]) == {"math", "pe"}
    assert updated_state["assignments"]["g1c1"]["chinese"] == []
    assert updated_state["assignments"]["g1c1"]["math"] == []
    assert updated_state["assignments"]["g1c1"]["reading"] == []
    assert _first_teacher_id(updated_state["assignments"]["g1c2"]["math"]) == teacher_id
    assert updated_state["assignments"]["g1c2"]["chinese"] == []
    assert updated_state["assignments"]["g1c2"]["reading"] == []
    assert _first_teacher_id(updated_state["assignments"]["g1c3"]["pe"]) == teacher_id

    before_invalid = local_store.load()
    invalid = client.put(
        f"/api/teachers/{teacher_id}",
        json={
            "name": "不应保存的新名称",
            "subject_ids": [],
            "min_weekly_lessons": 0,
            "teaching_assignments": {"g1c1": ["reading"]},
        },
    )
    assert invalid.status_code == 422
    assert local_store.load() == before_invalid


def test_teachers_can_be_deleted_in_one_atomic_batch(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    state = local_store.load()
    state["classes"] = [state["classes"][0]]
    state["assignments"] = {"g1c1": state["assignments"]["g1c1"]}
    state["teachers"] = [
        {"id": "t_delete_a", "name": "待删除甲", "subject_ids": ["chinese"], "min_weekly_lessons": 0, "homeroom_class_id": "g1c1"},
        {"id": "t_delete_b", "name": "待删除乙", "subject_ids": ["pe"], "min_weekly_lessons": 0, "homeroom_class_id": None},
        {"id": "t_keep", "name": "保留教师", "subject_ids": ["pe"], "min_weekly_lessons": 0, "homeroom_class_id": None},
    ]
    state["assignments"]["g1c1"]["chinese"] = "t_delete_a"
    state["assignments"]["g1c1"]["pe"] = [
        {"teacher_id": "t_delete_b", "lessons": 2},
        {"teacher_id": "t_keep", "lessons": 2},
    ]
    apply_required_teacher_assignments(state)
    state["schedule"] = {"success": True, "lessons": {}}
    local_store.save(state)
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)

    deleted = client.post(
        "/api/teachers/bulk-delete",
        json={"teacher_ids": ["t_delete_a", "t_delete_b"]},
    )
    assert deleted.status_code == 200, deleted.text
    assert deleted.json()["deleted"] == 2
    saved = deleted.json()["state"]
    assert [teacher["id"] for teacher in saved["teachers"]] == ["t_keep"]
    assert saved["schedule"] is None
    assert saved["assignments"]["g1c1"]["chinese"] == []
    assert saved["assignments"]["g1c1"]["reading"] == []
    assert saved["assignments"]["g1c1"]["meeting"] == []
    assert saved["assignments"]["g1c1"]["pe"] == [{"teacher_id": "t_keep", "lessons": 2}]

    before_missing = local_store.load()
    rejected = client.post(
        "/api/teachers/bulk-delete",
        json={"teacher_ids": ["t_keep", "missing-teacher"]},
    )
    assert rejected.status_code == 404
    assert local_store.load() == before_missing
    assert client.post("/api/teachers/bulk-delete", json={"teacher_ids": []}).status_code == 422


def test_batch_assignments_are_saved_atomically(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)
    _login(client)

    state = client.get("/api/state").json()["state"]
    first_class, second_class = state["classes"][:2]
    created = client.post(
        "/api/teachers",
        json={"name": "批量保存测试教师", "subject_ids": ["math"], "min_weekly_lessons": 0},
    )
    teacher_id = created.json()["state"]["teachers"][0]["id"]

    saved = client.put(
        "/api/assignments",
        json={
            "classes": {
                first_class["id"]: {"math": teacher_id},
                second_class["id"]: {"math": teacher_id},
            }
        },
    )
    assert saved.status_code == 200
    assert saved.json()["updated_classes"] == 2
    assert _first_teacher_id(saved.json()["state"]["assignments"][first_class["id"]]["math"]) == teacher_id
    assert _first_teacher_id(saved.json()["state"]["assignments"][second_class["id"]]["math"]) == teacher_id

    rejected = client.put(
        "/api/assignments",
        json={
            "classes": {
                first_class["id"]: {"math": None},
                "missing-class": {"math": None},
            }
        },
    )
    assert rejected.status_code == 404
    unchanged = client.get("/api/state").json()["state"]
    assert _first_teacher_id(unchanged["assignments"][first_class["id"]]["math"]) == teacher_id


def test_homeroom_and_chinese_teacher_drive_fixed_courses(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)
    _login(client)

    created = client.post(
        "/api/teachers",
        json={
            "name": "一班语文班主任",
            "subject_ids": ["chinese"],
            "min_weekly_lessons": 0,
            "homeroom_class_id": "g1c1",
        },
    )
    assert created.status_code == 200
    teacher_id = created.json()["state"]["teachers"][0]["id"]
    assert _first_teacher_id(created.json()["state"]["assignments"]["g1c1"]["meeting"]) == teacher_id

    assigned = client.put("/api/assignments/g1c1", json={"assignments": {"chinese": teacher_id}})
    assert assigned.status_code == 200
    assert _first_teacher_id(assigned.json()["state"]["assignments"]["g1c1"]["reading"]) == teacher_id

    conflict = client.post(
        "/api/teachers",
        json={
            "name": "冲突班主任",
            "subject_ids": [],
            "min_weekly_lessons": 0,
            "homeroom_class_id": "g1c1",
        },
    )
    assert conflict.status_code == 409


def test_teacher_template_and_batch_import(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)
    _login(client)

    template = client.get("/api/teachers/import-template.xlsx")
    assert template.status_code == 200
    template_workbook = load_workbook(BytesIO(template.content))
    assert template_workbook.sheetnames == ["教师名单", "填写说明"]
    assert [template_workbook["教师名单"].cell(1, column).value for column in range(1, 6)] == [
        "教师姓名\n（同名行合并）",
        "期望最低周课时\n（0—35整数；同名行相加）",
        "是否班主任\n（填“是”或“否”）",
        "班主任班级\n（推荐完整班名；简写按文本填写）",
        "任课配置\n（班级/范围：科目(节数)；见填写说明）",
    ]
    assert all(
        cell.fill.fill_type is None
        for sheet in template_workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert template_workbook["教师名单"]["A2"].border.right.style == "thin"
    assert template_workbook["教师名单"]["E2"].border.right.style == "thin"
    assert template_workbook["教师名单"]["D2"].number_format == "@"
    assert template_workbook["教师名单"]["E2"].number_format == "@"
    assert {str(item.sqref) for item in template_workbook["教师名单"].data_validations.dataValidation} == {
        "B2:B1001",
        "C2:C1001",
    }
    instructions = "\n".join(
        str(template_workbook["填写说明"].cell(row, 2).value or "")
        for row in range(2, 10)
    )
    assert "同一文件中姓名完全相同的多行自动合并为一位教师" in instructions
    assert "真实重名教师必须主动区分姓名" in instructions
    assert "同名三行分别填13、1、1，导入后该教师为15节" in instructions
    assert "三年级10班" in instructions
    assert "3.10" in instructions
    assert "自动改成数值“3.1”" in instructions
    assert "主授" not in instructions
    assert template_workbook["填写说明"]["A9"].border.right.style == "thin"

    created = client.post("/api/teachers", json={"name": "王老师", "subject_ids": ["english"], "min_weekly_lessons": 8})
    assert created.status_code == 200
    template_sheet = template_workbook["教师名单"]
    for row_index, values in enumerate((
        ["王老师", 12, "是", "1.1", "一年级1-2班：语文、数学"],
        ["王老师", 10, "", "", "1.1：体育"],
        ["李老师", 0, "否", "", ""],
    ), start=2):
        for column_index, value in enumerate(values, start=1):
            template_sheet.cell(row_index, column_index, value)
    template_output = BytesIO()
    template_workbook.save(template_output)
    content = template_output.getvalue()
    imported = client.post(
        "/api/teachers/import",
        files={"file": ("teachers.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["import"] | {"details": [], "shortages": []} == {
        "created": 1,
        "updated": 1,
        "merged_rows": 1,
        "source_rows": 3,
        "total": 2,
        "assigned": 5,
        "homerooms": 1,
        "shortages": [],
        "details": [],
    }
    teachers = {item["name"]: item for item in imported.json()["state"]["teachers"]}
    assert set(teachers) == {"王老师", "李老师"}
    assert teachers["王老师"]["subject_ids"] == ["chinese", "math", "pe"]
    assert teachers["王老师"]["min_weekly_lessons"] == 22
    assert teachers["李老师"]["subject_ids"] == []
    assert teachers["王老师"]["homeroom_class_id"] == "g1c1"
    imported_state = imported.json()["state"]
    assert _first_teacher_id(imported_state["assignments"]["g1c1"]["chinese"]) == teachers["王老师"]["id"]
    assert _first_teacher_id(imported_state["assignments"]["g1c1"]["reading"]) == teachers["王老师"]["id"]
    assert _first_teacher_id(imported_state["assignments"]["g1c1"]["meeting"]) == teachers["王老师"]["id"]
    assert _first_teacher_id(imported_state["assignments"]["g1c1"]["pe"]) == teachers["王老师"]["id"]
    assert _first_teacher_id(imported_state["assignments"]["g1c2"]["chinese"]) == teachers["王老师"]["id"]
    assert _first_teacher_id(imported_state["assignments"]["g1c2"]["math"]) == teachers["王老师"]["id"]

    imported_again = client.post(
        "/api/teachers/import",
        files={"file": ("teachers.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported_again.status_code == 200
    assert imported_again.json()["import"]["created"] == 0
    assert imported_again.json()["import"]["updated"] == 2
    assert len(imported_again.json()["state"]["teachers"]) == 2

    bad_content = _teacher_import_file([["错误教师", 12, "否", "", "一年级1班：不存在的科目"]])
    rejected = client.post("/api/teachers/import", files={"file": ("bad.xlsx", bad_content)})
    assert rejected.status_code == 422
    assert len(client.get("/api/state").json()["state"]["teachers"]) == 2

    bad_range = _teacher_import_file([["范围错误教师", 12, "否", "", "一年级1-7班：语文"]])
    rejected_range = client.post("/api/teachers/import", files={"file": ("bad-range.xlsx", bad_range)})
    assert rejected_range.status_code == 422
    assert "一年级7班" in rejected_range.json()["detail"]


def test_teacher_import_merges_repeated_name_rows_and_requires_manual_homonym_labels(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)
    _login(client)

    repeated = _teacher_import_file([
        ["郭传荣", 13, "是", "2.1", "二年级1班：语文"],
        ["郭传荣", 1, "", "", "二年级1班：校本课程"],
        ["郭传荣", 1, "", "", "二年级1班：实践"],
        ["张伟（语文）", 8, "否", "", "二年级2班：语文"],
        ["张伟（体育）", 4, "否", "", "二年级2班：体育"],
    ])
    imported = client.post("/api/teachers/import", files={"file": ("repeated.xlsx", repeated)})
    assert imported.status_code == 200, imported.text
    summary = imported.json()["import"]
    assert {key: summary[key] for key in ("source_rows", "total", "merged_rows", "created", "updated")} == {
        "source_rows": 5,
        "total": 3,
        "merged_rows": 2,
        "created": 3,
        "updated": 0,
    }
    teachers = {item["name"]: item for item in imported.json()["state"]["teachers"]}
    assert set(teachers) == {"郭传荣", "张伟（语文）", "张伟（体育）"}
    assert teachers["郭传荣"]["min_weekly_lessons"] == 15
    assert teachers["郭传荣"]["subject_ids"] == ["chinese", "school", "practice"]
    assert teachers["郭传荣"]["homeroom_class_id"] == "g2c1"
    state = imported.json()["state"]
    for subject_id in ("chinese", "school", "practice"):
        assert _first_teacher_id(state["assignments"]["g2c1"][subject_id]) == teachers["郭传荣"]["id"]

    conflict = _teacher_import_file([
        ["同一教师", 8, "是", "1.1", "一年级1班：语文"],
        ["同一教师", 3, "是", "1.2", "一年级2班：数学"],
    ])
    rejected = client.post("/api/teachers/import", files={"file": ("conflict.xlsx", conflict)})
    assert rejected.status_code == 422
    assert "班主任信息不一致" in rejected.json()["detail"]


def test_teacher_import_distinguishes_tenth_class_and_rejects_excel_numeric_ambiguity(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    local_store.update_settings(
        "测试学校",
        {str(grade): (10 if grade == 3 else 6) for grade in range(1, 7)},
    )
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)

    exact = _teacher_import_file([
        ["十班班主任", 0, "是", "3.10", "3.9-3.10：语文"],
    ])
    imported = client.post("/api/teachers/import", files={"file": ("exact.xlsx", exact)})
    assert imported.status_code == 200, imported.text
    teacher = next(item for item in imported.json()["state"]["teachers"] if item["name"] == "十班班主任")
    assert teacher["homeroom_class_id"] == "g3c10"
    assert _first_teacher_id(imported.json()["state"]["assignments"]["g3c9"]["chinese"]) == teacher["id"]
    assert _first_teacher_id(imported.json()["state"]["assignments"]["g3c10"]["chinese"]) == teacher["id"]

    ambiguous = _teacher_import_file([
        ["数值简写教师", 0, "是", 3.10, "三年级10班：数学"],
    ])
    rejected = client.post("/api/teachers/import", files={"file": ("ambiguous.xlsx", ambiguous)})
    assert rejected.status_code == 422
    assert "被Excel保存为数值3.1" in rejected.json()["detail"]
    assert "三年级10班" in rejected.json()["detail"]
    assert all(item["name"] != "数值简写教师" for item in client.get("/api/state").json()["state"]["teachers"])


def test_teacher_import_rejects_duplicate_custom_class_names(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(main, "store", StateStore(tmp_path / "state.json"))
    client = TestClient(main.app)
    _login(client)
    assert client.put("/api/classes/g3c1", json={"name": "同名班"}).status_code == 200
    assert client.put("/api/classes/g3c2", json={"name": "同名班"}).status_code == 200

    content = _teacher_import_file([
        ["名称歧义教师", 0, "是", "同名班", "三年级1班：语文"],
    ])
    rejected = client.post("/api/teachers/import", files={"file": ("duplicate-name.xlsx", content)})
    assert rejected.status_code == 422
    assert "班主任班级名称不唯一" in rejected.json()["detail"]
    assert "g3c1、g3c2" in rejected.json()["detail"]


def test_teacher_import_supports_shared_lesson_quotas_and_strong_shortage_confirmation(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)

    exact_content = _teacher_import_file([
        ["体育甲", 0, "否", "", "一年级1班：体育(2节)"],
        ["体育乙", 0, "否", "", "一年级1班：体育(1节)"],
        ["体育丙", 0, "否", "", "一年级1班：体育(1节)"],
    ])
    imported = client.post("/api/teachers/import", files={"file": ("exact.xlsx", exact_content)})
    assert imported.status_code == 200, imported.text
    assert not imported.json().get("requires_confirmation")
    teachers = {item["name"]: item["id"] for item in imported.json()["state"]["teachers"]}
    pe_allocations = imported.json()["state"]["assignments"]["g1c1"]["pe"]
    assert [(item["teacher_id"], item["lessons"]) for item in pe_allocations] == [
        (teachers["体育甲"], 2),
        (teachers["体育乙"], 1),
        (teachers["体育丙"], 1),
    ]
    assert all(set(item) == {"teacher_id", "lessons"} for item in pe_allocations)

    shortage_content = _teacher_import_file([
        ["二班体育", 0, "否", "", "一年级2班：体育(2节)"],
    ])
    preview = client.post("/api/teachers/import", files={"file": ("short.xlsx", shortage_content)})
    assert preview.status_code == 200
    assert preview.json()["requires_confirmation"] is True
    shortage = preview.json()["shortages"][0]
    assert {key: shortage[key] for key in ("class_id", "assigned", "expected", "missing")} == {
        "class_id": "g1c2",
        "assigned": 2,
        "expected": 4,
        "missing": 2,
    }
    assert "二班体育" not in {item["name"] for item in client.get("/api/state").json()["state"]["teachers"]}

    confirmed = client.post(
        "/api/teachers/import?allow_incomplete=true",
        files={"file": ("short.xlsx", shortage_content)},
    )
    assert confirmed.status_code == 200
    assert "二班体育" in {item["name"] for item in confirmed.json()["state"]["teachers"]}

    over_content = _teacher_import_file([
        ["三班体育甲", 0, "否", "", "一年级3班：体育(3节)"],
        ["三班体育乙", 0, "否", "", "一年级3班：体育(2节)"],
    ])
    before_names = {item["name"] for item in client.get("/api/state").json()["state"]["teachers"]}
    rejected = client.post("/api/teachers/import", files={"file": ("over.xlsx", over_content)})
    assert rejected.status_code == 422
    assert "超过标准4节" in rejected.json()["detail"]
    assert {item["name"] for item in client.get("/api/state").json()["state"]["teachers"]} == before_names


def test_teacher_schedule_export_by_grade_or_teacher(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    scoped_state = local_store.load()
    scoped_state["classes"] = [item for item in scoped_state["classes"] if item["id"] == "g3c1"]
    scoped_state["assignments"] = {"g3c1": scoped_state["assignments"]["g3c1"]}
    local_store.save(scoped_state)
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)
    teacher = client.post(
        "/api/teachers",
        json={"name": "三年级语文教师", "subject_ids": ["chinese"], "min_weekly_lessons": 0, "homeroom_class_id": "g3c1"},
    ).json()["state"]["teachers"][0]
    assigned = client.put("/api/assignments/g3c1", json={"assignments": {"chinese": teacher["id"]}})
    assert assigned.status_code == 200
    generated = _generate_and_wait(client, {"seed": 818, "attempts": 40})
    assert generated["job"]["status"] == "completed"
    assert generated["job"]["result"]["success"]

    for url in (
        "/api/export/teachers.zip",
        "/api/export/teachers.zip?grade=3",
    ):
        exported = client.get(url)
        assert exported.status_code == 200, exported.text
        with ZipFile(BytesIO(exported.content)) as archive:
            names = archive.namelist()
            assert names == ["三年级语文教师课程表.pdf"]
            teacher_pdf = archive.read(names[0])
            assert teacher_pdf.startswith(b"%PDF")
            assert len(PdfReader(BytesIO(teacher_pdf)).pages) == 1

    single_export = client.get(f"/api/export/teacher.pdf?teacher_id={teacher['id']}")
    assert single_export.status_code == 200
    assert single_export.content.startswith(b"%PDF")
    assert len(single_export.content) < 500_000
    assert single_export.headers["content-type"] == "application/pdf"
    assert single_export.headers["content-disposition"].endswith(".pdf")
    single_pdf = PdfReader(BytesIO(single_export.content))
    assert len(single_pdf.pages) == 1
    single_pdf_text = single_pdf.pages[0].extract_text() or ""
    assert "高唐县民族实验小学教师课程表" in single_pdf_text
    assert "教师：三年级语文教师" in single_pdf_text
    assert float(single_pdf.pages[0].mediabox.height) > float(single_pdf.pages[0].mediabox.width)
    assert _page_contains_image(single_pdf.pages[0])

    for url in (
        "/api/export/teachers.xlsx",
        "/api/export/teachers.xlsx?grade=3",
        f"/api/export/teachers.xlsx?teacher_id={teacher['id']}",
    ):
        exported_xlsx = client.get(url)
        assert exported_xlsx.status_code == 200, exported_xlsx.text
        teacher_workbook = load_workbook(BytesIO(exported_xlsx.content))
        assert teacher_workbook.sheetnames == ["三年级语文教师"]
        teacher_sheet = teacher_workbook.active
        assert teacher_sheet["A1"].value == "高唐县民族实验小学教师课程表"
        assert teacher_sheet["A2"].value == "教师：三年级语文教师"
        assert teacher_sheet.page_setup.orientation == "portrait"
        assert teacher_sheet.print_area == "'三年级语文教师'!$A$1:$G$10"
        assert teacher_sheet.print_title_rows == "$1:$3"
        assert teacher_sheet.row_dimensions[4].height == 92
        assert all(cell.fill.fill_type is None for row in teacher_sheet.iter_rows() for cell in row)

    assert client.get("/api/export/teachers.zip?grade=1").status_code == 409
    assert client.get("/api/export/teacher.pdf?teacher_id=missing").status_code == 404
    assert client.get("/api/export/teachers.xlsx?teacher_id=missing").status_code == 404


def test_schedule_swaps_are_atomic_and_keep_teacher_views_consistent(tmp_path, monkeypatch) -> None:
    local_store = StateStore(tmp_path / "state.json")
    scoped_state = local_store.load()
    scoped_state["classes"] = [item for item in scoped_state["classes"] if item["id"] in {"g3c1", "g3c2"}]
    scoped_state["assignments"] = {
        class_id: scoped_state["assignments"][class_id]
        for class_id in ("g3c1", "g3c2")
    }
    local_store.save(scoped_state)
    monkeypatch.setattr(main, "store", local_store)
    client = TestClient(main.app)
    _login(client)

    generated = _generate_and_wait(client, {"seed": 20260822, "attempts": 20})
    assert generated["job"]["status"] == "completed"
    assert generated["job"]["result"]["success"]

    state = local_store.load()
    first_subject = state["schedule"]["lessons"]["g3c1"]["mon-am3"]["subject_id"]
    second_subject = state["schedule"]["lessons"]["g3c2"]["tue-am3"]["subject_id"]
    state["teachers"].append({
        "id": "t_manual",
        "name": "手动调整测试教师",
        "subject_ids": list(dict.fromkeys([first_subject, second_subject])),
        "min_weekly_lessons": 0,
        "homeroom_class_id": None,
    })
    state["schedule"]["lessons"]["g3c1"]["mon-am3"]["teacher_id"] = "t_manual"
    state["schedule"]["lessons"]["g3c2"]["tue-am3"]["teacher_id"] = "t_manual"
    local_store.save(state)
    before_lessons = local_store.load()["schedule"]["lessons"]

    fixed_rejected = client.put(
        "/api/schedule/swaps",
        json={"class_id": "g3c1", "swaps": [{"from_slot": "mon-pm3", "to_slot": "mon-am3"}]},
    )
    assert fixed_rejected.status_code == 409
    assert "固定课位" in "；".join(fixed_rejected.json()["detail"])
    assert local_store.load()["schedule"]["lessons"] == before_lessons

    conflict = client.put(
        "/api/schedule/swaps",
        json={"class_id": "g3c1", "swaps": [{"from_slot": "mon-am3", "to_slot": "tue-am3"}]},
    )
    assert conflict.status_code == 409
    assert "手动调整测试教师" in "；".join(conflict.json()["detail"])
    assert local_store.load()["schedule"]["lessons"] == before_lessons

    saved = client.put(
        "/api/schedule/swaps",
        json={
            "class_id": "g3c1",
            "swaps": [
                {"from_slot": "mon-am3", "to_slot": "wed-am3"},
                {"from_slot": "wed-am3", "to_slot": "fri-am3"},
            ],
        },
    )
    assert saved.status_code == 200, saved.text
    schedule = saved.json()["state"]["schedule"]
    assert saved.json()["updated_swaps"] == 2
    assert schedule["manual_swap_count"] == 2
    assert schedule["edited_at"]
    assert schedule["lessons"]["g3c1"]["fri-am3"]["teacher_id"] == "t_manual"
    assert schedule["lessons"]["g3c1"]["fri-am3"]["subject_id"] == first_subject
    assert schedule["lessons"]["g3c1"]["mon-am3"]["teacher_id"] is None

    teacher_xlsx = client.get("/api/export/teachers.xlsx?teacher_id=t_manual")
    assert teacher_xlsx.status_code == 200
    teacher_sheet = load_workbook(BytesIO(teacher_xlsx.content)).active
    assert str(teacher_sheet["G6"].value).startswith("三年级1班\n")


def test_upstash_store_initializes_and_persists(monkeypatch) -> None:
    remote: dict[str, str] = {}
    cloud_store = UpstashStateStore("https://example.upstash.io", "test-token", key="test-state")

    def fake_execute(command: str, *arguments: object) -> object:
        if command == "GET":
            return remote.get(str(arguments[0]))
        if command == "SET":
            remote[str(arguments[0])] = str(arguments[1])
            return "OK"
        raise AssertionError(command)

    monkeypatch.setattr(cloud_store, "_execute", fake_execute)
    initial = cloud_store.load()
    assert initial["school_name"] == "高唐县民族实验小学"
    assert json.loads(remote["test-state"])["schema_version"] == 5

    initial["school_name"] = "云端测试学校"
    cloud_store.save(initial)
    assert cloud_store.load()["school_name"] == "云端测试学校"

    def failed_execute(command: str, *arguments: object) -> object:
        raise StorageError("云端测试连接失败")

    monkeypatch.setattr(cloud_store, "_execute", failed_execute)
    monkeypatch.setattr(main, "store", cloud_store)
    client = TestClient(main.app)
    _login(client)
    unavailable = client.get("/api/state")
    assert unavailable.status_code == 503
    assert unavailable.json()["detail"] == "云端测试连接失败"
