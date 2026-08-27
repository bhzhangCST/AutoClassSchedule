from __future__ import annotations

import re
import uuid
from io import BytesIO
from itertools import islice
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .assignment_rules import apply_required_teacher_assignments
from .constants import CURRICULUM, GRADE_NAMES, SUBJECTS
from .teaching_allocations import (
    allocation_teacher_ids,
    allocation_total,
    normalize_allocations,
    remove_teacher,
)

MAX_IMPORT_ROWS = 1000
NAME_HEADERS = {"教师姓名", "姓名", "老师姓名"}
LESSON_HEADERS = {"期望最低周课时", "课时数", "最低周课时", "周课时"}
HOMEROOM_FLAG_HEADERS = {"是否班主任", "班主任标记"}
HOMEROOM_CLASS_HEADERS = {"班主任班级", "所带班级"}
ASSIGNMENT_HEADERS = {"任课配置", "任教班级与科目", "班级任课"}
LOCKED_SUBJECTS = {"reading": "阅读", "meeting": "班队会"}
CHINESE_GRADE_NUMBERS = {name.removesuffix("年级"): grade for grade, name in GRADE_NAMES.items()}
RANGE_SEPARATOR = r"[-—–至到~～]"


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s（）()：:]", "", str(value or "").strip())


def _header_index(headers: list[Any], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header(item) for item in candidates}
    for index, value in enumerate(headers):
        normalized = _normalize_header(value)
        if normalized in normalized_candidates:
            return index
        if any(normalized.startswith(candidate) for candidate in normalized_candidates):
            return index
    return None


def _subject_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for subject in SUBJECTS:
        for value in (subject["id"], subject["name"], subject["short"]):
            aliases[_normalize_header(value).lower()] = subject["id"]
    aliases.update({
        "道德法治": "morality",
        "体育健康": "pe",
        "信息技术": "it",
        "班队会": "meeting",
        "安全课": "meeting",
    })
    return aliases


SUBJECT_ALIASES = _subject_aliases()


def _parse_subject_specs(value: Any, row_number: int) -> list[dict[str, Any]]:
    text = str(value or "").strip()
    if not text:
        return []
    tokens = [item.strip() for item in re.split(r"[、,，/|]+", text) if item.strip()]
    specs: list[dict[str, Any]] = []
    for token in tokens:
        lesson_match = re.search(r"[（(]\s*(\d+)\s*节\s*[）)]$", token)
        subject_text = token[: lesson_match.start()].strip() if lesson_match else token
        subject_id = SUBJECT_ALIASES.get(_normalize_header(subject_text).lower())
        if not subject_id:
            raise ValueError(f"第{row_number}行包含未知科目：{subject_text}")
        lessons = int(lesson_match.group(1)) if lesson_match else None
        if lessons is not None and lessons <= 0:
            raise ValueError(f"第{row_number}行科目课时必须大于0：{token}")
        specs.append({"subject_id": subject_id, "lessons": lessons})
    return specs


def _parse_homeroom(flag: Any, class_name: Any, row_number: int) -> tuple[bool, str | None]:
    flag_text = str(flag or "").strip().lower()
    class_text = str(class_name or "").strip()
    if not flag_text and not class_text:
        return False, None
    if flag_text in {"是", "有", "1", "true", "yes", "y"}:
        if not class_text:
            raise ValueError(f"第{row_number}行选择了班主任，但未填写班主任班级")
        return True, class_text
    if flag_text in {"否", "无", "0", "false", "no", "n"}:
        if class_text:
            raise ValueError(f"第{row_number}行填写为非班主任，不应再填写班主任班级")
        return True, None
    if not flag_text and class_text:
        return True, class_text
    raise ValueError(f"第{row_number}行“是否班主任”只能填写是或否")


def _parse_assignment_entries(value: Any, row_number: int) -> list[dict[str, Any]]:
    text = str(value or "").strip()
    if not text:
        return []
    entries: list[dict[str, Any]] = []
    for segment in [item.strip() for item in re.split(r"[；;\n]+", text) if item.strip()]:
        parts = re.split(r"[：:]", segment, maxsplit=1)
        if len(parts) != 2 or not parts[0].strip() or not parts[1].strip():
            raise ValueError(f"第{row_number}行任课配置格式错误：{segment}")
        subject_specs = _parse_subject_specs(parts[1], row_number)
        if not subject_specs:
            raise ValueError(f"第{row_number}行任课配置“{segment}”没有有效科目")
        locked = [
            LOCKED_SUBJECTS[item["subject_id"]]
            for item in subject_specs
            if item["subject_id"] in LOCKED_SUBJECTS
        ]
        if locked:
            raise ValueError(
                f"第{row_number}行无需单独配置{'、'.join(locked)}；"
                "阅读随语文教师，班队会随班主任自动设置"
            )
        entries.append({"class_name": parts[0].strip(), "subject_specs": subject_specs})
    return entries


def _grade_number(value: str) -> int | None:
    text = value.strip()
    if text.isdigit() and 1 <= int(text) <= 6:
        return int(text)
    return CHINESE_GRADE_NUMBERS.get(text)


def _class_id(grade: int, number: int) -> str:
    return f"g{grade}c{number}"


def _resolve_class_reference(
    value: str,
    classes_by_name: dict[str, dict[str, Any]],
    classes_by_id: dict[str, dict[str, Any]],
    *,
    row_number: int,
    field_name: str,
    allow_range: bool,
) -> list[dict[str, Any]]:
    text = re.sub(r"\s+", "", str(value or ""))
    direct = classes_by_name.get(text) or classes_by_id.get(text)
    if direct:
        return [direct]

    exact_match = re.fullmatch(r"([一二三四五六1-6])年级(\d+)班", text)
    if not exact_match:
        exact_match = re.fullmatch(r"([1-6])[.．](\d+)", text)
    if exact_match:
        grade = _grade_number(exact_match.group(1))
        class_item = classes_by_id.get(_class_id(grade, int(exact_match.group(2)))) if grade else None
        if class_item:
            return [class_item]
        raise ValueError(f"第{row_number}行{field_name}不存在：{value}")

    if allow_range:
        range_match = re.fullmatch(
            rf"([一二三四五六1-6])年级(\d+)班?{RANGE_SEPARATOR}(\d+)班",
            text,
        )
        numeric_range_match = re.fullmatch(
            rf"([1-6])[.．](\d+){RANGE_SEPARATOR}(?:([1-6])[.．])?(\d+)",
            text,
        )
        if range_match:
            grade = _grade_number(range_match.group(1))
            start, end = int(range_match.group(2)), int(range_match.group(3))
        elif numeric_range_match:
            grade = int(numeric_range_match.group(1))
            end_grade = int(numeric_range_match.group(3)) if numeric_range_match.group(3) else grade
            if end_grade != grade:
                raise ValueError(f"第{row_number}行{field_name}范围不能跨年级：{value}")
            start, end = int(numeric_range_match.group(2)), int(numeric_range_match.group(4))
        else:
            grade = start = end = None
        if grade is not None:
            if start > end:
                raise ValueError(f"第{row_number}行{field_name}范围起始班级不能大于结束班级：{value}")
            resolved: list[dict[str, Any]] = []
            missing: list[str] = []
            for number in range(start, end + 1):
                class_item = classes_by_id.get(_class_id(grade, number))
                if class_item:
                    resolved.append(class_item)
                else:
                    missing.append(f"{GRADE_NAMES[grade]}{number}班")
            if missing:
                raise ValueError(f"第{row_number}行{field_name}范围包含不存在的班级：{'、'.join(missing)}")
            return resolved

    supported = "完整班名（如一年级1班）或数字简写（如1.1）"
    if allow_range:
        supported += "，范围可写一年级1-7班或1.1-1.7"
    raise ValueError(f"第{row_number}行{field_name}格式无法识别：{value}；请填写{supported}")


def parse_teacher_workbook(content: bytes) -> list[dict[str, Any]]:
    if not content:
        raise ValueError("上传文件为空")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise ValueError("无法读取该文件，请上传有效的 .xlsx 教师名单") from exc

    sheet = workbook["教师名单"] if "教师名单" in workbook.sheetnames else workbook.active
    rows = list(islice(sheet.iter_rows(values_only=True), MAX_IMPORT_ROWS + 2))
    if not rows:
        raise ValueError("教师名单中没有表头")
    headers = list(rows[0])
    name_index = _header_index(headers, NAME_HEADERS)
    lesson_index = _header_index(headers, LESSON_HEADERS)
    homeroom_flag_index = _header_index(headers, HOMEROOM_FLAG_HEADERS)
    homeroom_class_index = _header_index(headers, HOMEROOM_CLASS_HEADERS)
    assignment_index = _header_index(headers, ASSIGNMENT_HEADERS)
    required_indexes = (name_index, lesson_index, assignment_index)
    missing = [
        label for label, index in zip(("教师姓名", "期望最低周课时", "任课配置"), required_indexes)
        if index is None
    ]
    if missing:
        raise ValueError(f"缺少表头：{'、'.join(missing)}，请使用下载模板填写")

    present_indexes = [index for index in (
        name_index, lesson_index, homeroom_flag_index, homeroom_class_index, assignment_index
    ) if index is not None]
    parsed: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_number, row in enumerate(rows[1 : MAX_IMPORT_ROWS + 1], start=2):
        values = list(row)
        if not any(index < len(values) and values[index] is not None and str(values[index]).strip() for index in present_indexes):
            continue
        name = str(values[name_index] or "").strip() if name_index is not None and name_index < len(values) else ""
        if not name:
            errors.append(f"第{row_number}行缺少教师姓名")
            continue
        if len(name) > 40:
            errors.append(f"第{row_number}行教师姓名超过40个字符")
            continue
        raw_lessons = values[lesson_index] if lesson_index is not None and lesson_index < len(values) else None
        try:
            if raw_lessons is None or str(raw_lessons).strip() == "":
                raise ValueError
            lessons_number = float(raw_lessons)
            if not lessons_number.is_integer():
                raise ValueError
            min_weekly_lessons = int(lessons_number)
            if not 0 <= min_weekly_lessons <= 35:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"第{row_number}行课时数必须是0—35的整数")
            continue
        try:
            homeroom_specified, homeroom_class_name = _parse_homeroom(
                values[homeroom_flag_index] if homeroom_flag_index is not None and homeroom_flag_index < len(values) else None,
                values[homeroom_class_index] if homeroom_class_index is not None and homeroom_class_index < len(values) else None,
                row_number,
            )
            assignment_entries = _parse_assignment_entries(
                values[assignment_index] if assignment_index is not None and assignment_index < len(values) else None,
                row_number,
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue
        parsed.append({
            "row": row_number,
            "name": name,
            "min_weekly_lessons": min_weekly_lessons,
            "homeroom_specified": homeroom_specified,
            "homeroom_class_name": homeroom_class_name,
            "assignment_entries": assignment_entries,
        })

    if len(rows) > MAX_IMPORT_ROWS + 1:
        errors.append(f"单次最多导入{MAX_IMPORT_ROWS}行教师数据")
    if errors:
        shown = errors[:12]
        suffix = f"；另有{len(errors) - len(shown)}项错误" if len(errors) > len(shown) else ""
        raise ValueError("；".join(shown) + suffix)
    if not parsed:
        raise ValueError("教师名单中没有可导入的数据")
    return parsed


def merge_teacher_rows(state: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
    teachers = state.setdefault("teachers", [])
    classes = state.get("classes", [])
    classes_by_name = {item["name"].strip(): item for item in classes}
    classes_by_id = {item["id"]: item for item in classes}
    assignment_claims: dict[tuple[str, str], list[dict[str, Any]]] = {}
    homeroom_claims: dict[str, str] = {}
    prepared_by_name: dict[str, dict[str, Any]] = {}

    for row in rows:
        teacher_name = row["name"]
        prepared = prepared_by_name.setdefault(teacher_name, {
            "name": teacher_name,
            "row": row["row"],
            "source_rows": [],
            "min_weekly_lessons": 0,
            "homeroom_specified": False,
            "homeroom_class_id": None,
            "subject_ids": [],
        })
        prepared["source_rows"].append(row["row"])
        prepared["min_weekly_lessons"] += int(row["min_weekly_lessons"])
        if prepared["min_weekly_lessons"] > 35:
            row_labels = "、".join(str(item) for item in prepared["source_rows"])
            raise ValueError(
                f"同名教师{teacher_name}在第{row_labels}行的期望最低周课时合计"
                f"{prepared['min_weekly_lessons']}节，超过35节"
            )

        homeroom_class_id = None
        if row.get("homeroom_class_name"):
            class_item = _resolve_class_reference(
                row["homeroom_class_name"],
                classes_by_name,
                classes_by_id,
                row_number=row["row"],
                field_name="班主任班级",
                allow_range=False,
            )[0]
            homeroom_class_id = class_item["id"]
            claimed_by = homeroom_claims.get(homeroom_class_id)
            if claimed_by and claimed_by != teacher_name:
                raise ValueError(f"{class_item['name']}在上传文件中被同时指定给{claimed_by}和{teacher_name}担任班主任")
            homeroom_claims[homeroom_class_id] = teacher_name
        if row.get("homeroom_specified"):
            if prepared["homeroom_specified"] and prepared["homeroom_class_id"] != homeroom_class_id:
                previous = (
                    classes_by_id[prepared["homeroom_class_id"]]["name"]
                    if prepared["homeroom_class_id"]
                    else "非班主任"
                )
                current = classes_by_id[homeroom_class_id]["name"] if homeroom_class_id else "非班主任"
                raise ValueError(
                    f"同名教师{teacher_name}的班主任信息不一致："
                    f"先前填写为{previous}，第{row['row']}行填写为{current}"
                )
            prepared["homeroom_specified"] = True
            prepared["homeroom_class_id"] = homeroom_class_id

        for entry in row.get("assignment_entries", []):
            class_items = _resolve_class_reference(
                entry["class_name"],
                classes_by_name,
                classes_by_id,
                row_number=row["row"],
                field_name="任课班级",
                allow_range=True,
            )
            for class_item in class_items:
                allowed = set(CURRICULUM[int(class_item["grade"])])
                for subject_spec in entry["subject_specs"]:
                    subject_id = subject_spec["subject_id"]
                    if subject_id not in allowed:
                        raise ValueError(f"第{row['row']}行：{class_item['name']}没有该科目")
                    key = (class_item["id"], subject_id)
                    expected = CURRICULUM[int(class_item["grade"])][subject_id]
                    lessons = subject_spec["lessons"] if subject_spec["lessons"] is not None else expected
                    if lessons > expected:
                        subject_name = next(item["name"] for item in SUBJECTS if item["id"] == subject_id)
                        raise ValueError(
                            f"第{row['row']}行：{class_item['name']}的{subject_name}填写{lessons}节，超过标准{expected}节"
                        )
                    assignment_claims.setdefault(key, []).append({
                        "teacher_name": teacher_name,
                        "lessons": lessons,
                        "row": row["row"],
                    })
                    if subject_id not in prepared["subject_ids"]:
                        prepared["subject_ids"].append(subject_id)

    prepared_rows = list(prepared_by_name.values())

    by_name = {teacher["name"]: teacher for teacher in teachers}
    details: list[dict[str, Any]] = []
    targets: dict[str, dict[str, Any]] = {}
    created = updated = 0
    for row in prepared_rows:
        target = by_name.get(row["name"])
        if target:
            target.update({
                "subject_ids": list(row["subject_ids"]),
                "min_weekly_lessons": int(row["min_weekly_lessons"]),
            })
            action = "updated"
            updated += 1
        else:
            target = {
                "id": f"t_{uuid.uuid4().hex[:10]}",
                "name": row["name"],
                "subject_ids": list(row["subject_ids"]),
                "min_weekly_lessons": int(row["min_weekly_lessons"]),
                "homeroom_class_id": None,
            }
            teachers.append(target)
            by_name[row["name"]] = target
            action = "created"
            created += 1
        if row.get("homeroom_specified"):
            target["homeroom_class_id"] = row["homeroom_class_id"]
        targets[row["name"]] = target
        details.append({
            "row": row["row"],
            "source_name": row["name"],
            "source_rows": row["source_rows"],
            "teacher_name": row["name"],
            "action": action,
        })

    for class_id, teacher_name in homeroom_claims.items():
        target = targets[teacher_name]
        for teacher in teachers:
            if teacher["id"] != target["id"] and teacher.get("homeroom_class_id") == class_id:
                teacher["homeroom_class_id"] = None
    touched: set[tuple[str, str]] = set(assignment_claims)
    target_ids = {target["id"] for target in targets.values()}
    classes_by_id = {item["id"]: item for item in classes}
    for class_id, class_assignments in state["assignments"].items():
        grade = int(classes_by_id[class_id]["grade"])
        for subject_id, assigned_value in list(class_assignments.items()):
            if subject_id in LOCKED_SUBJECTS:
                continue
            original_ids = allocation_teacher_ids(assigned_value)
            if original_ids & target_ids:
                touched.add((class_id, subject_id))
            value = assigned_value
            for teacher_id in target_ids:
                value = remove_teacher(value, teacher_id, CURRICULUM[grade][subject_id])
            class_assignments[subject_id] = value

    for (class_id, subject_id), claims in assignment_claims.items():
        class_item = classes_by_id[class_id]
        expected = CURRICULUM[int(class_item["grade"])][subject_id]
        allocations = list(state["assignments"][class_id].get(subject_id, []))
        for claim in claims:
            teacher_id = targets[claim["teacher_name"]]["id"]
            existing = next(
                (item for item in allocations if item.get("teacher_id") == teacher_id),
                None,
            )
            if existing:
                existing["lessons"] += int(claim["lessons"])
            else:
                allocations.append({
                    "teacher_id": teacher_id,
                    "lessons": int(claim["lessons"]),
                })
        normalized = normalize_allocations(allocations, expected)
        assigned = allocation_total(normalized)
        if assigned > expected:
            subject_name = next(item["name"] for item in SUBJECTS if item["id"] == subject_id)
            teacher_parts = "、".join(
                f"{claim['teacher_name']}{claim['lessons']}节" for claim in claims
            )
            raise ValueError(
                f"{class_item['name']}的{subject_name}合计{assigned}节，超过标准{expected}节（本次：{teacher_parts}）"
            )
        state["assignments"][class_id][subject_id] = normalized

    shortages: list[dict[str, Any]] = []
    for class_id, subject_id in sorted(touched):
        class_item = classes_by_id[class_id]
        expected = CURRICULUM[int(class_item["grade"])][subject_id]
        assigned = allocation_total(state["assignments"][class_id][subject_id])
        if 0 < assigned < expected:
            subject_name = next(item["name"] for item in SUBJECTS if item["id"] == subject_id)
            shortages.append({
                "class_id": class_id,
                "class_name": class_item["name"],
                "subject_id": subject_id,
                "subject_name": subject_name,
                "assigned": assigned,
                "expected": expected,
                "missing": expected - assigned,
            })

    apply_required_teacher_assignments(state)
    state["schedule"] = None
    return {
        "created": created,
        "updated": updated,
        "merged_rows": len(rows) - len(prepared_rows),
        "source_rows": len(rows),
        "total": len(prepared_rows),
        "assigned": len(assignment_claims),
        "homerooms": len(homeroom_claims),
        "shortages": shortages,
        "details": details,
    }
