from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

from .constants import DAYS, PERIODS, SUBJECT_BY_ID, slots_for_grade

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

THIN_SIDE = Side(style="thin", color="666666")
MEDIUM_SIDE = Side(style="medium", color="333333")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TITLE_FONT = Font(name="Microsoft YaHei", size=18, bold=True, color="222222")
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, bold=False, color="333333")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="222222")
BODY_FONT = Font(name="SimSun", size=10, color="222222")


def _validated_schedule(state: dict[str, Any]) -> dict[str, Any]:
    schedule = state.get("schedule")
    if not schedule or not schedule.get("success"):
        raise ValueError("尚未生成可导出的课表")
    return schedule


def _safe_sheet_title(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", value).strip() or "课程表"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.lower() in used:
        suffix = f"-{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate.lower())
    return candidate


def _configure_sheet(
    sheet: Any,
    title: str,
    cell_value: Callable[[str], str],
    unavailable_slots: set[str] | None = None,
    subtitle: str | None = None,
) -> None:
    unavailable_slots = unavailable_slots or set()
    header_row = 3 if subtitle else 2
    first_period_row = header_row + 1
    last_period_row = first_period_row + len(PERIODS) - 1
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = f"C{first_period_row}"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 42

    if subtitle:
        sheet.merge_cells("A2:G2")
        sheet["A2"] = subtitle
        sheet["A2"].font = SUBTITLE_FONT
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[2].height = 22

    headers = ["时段", "节次", *[day["name"] for day in DAYS]]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=THIN_SIDE, right=THIN_SIDE, top=MEDIUM_SIDE, bottom=THIN_SIDE)
    sheet.row_dimensions[header_row].height = 28

    for period_offset, period in enumerate(PERIODS):
        period_row = first_period_row + period_offset
        sheet.cell(period_row, 1, period["section"] if period_offset in (0, 4) else "")
        sheet.cell(period_row, 2, f"第{period['order']}节")
        for day_index, day in enumerate(DAYS, start=3):
            slot = f"{day['id']}-{period['id']}"
            sheet.cell(period_row, day_index, "—" if slot in unavailable_slots else cell_value(slot))
        for column in range(1, 8):
            cell = sheet.cell(period_row, column)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = TABLE_BORDER
        sheet.row_dimensions[period_row].height = 92

    morning_cell = f"A{first_period_row}"
    afternoon_cell = f"A{first_period_row + 4}"
    sheet.merge_cells(start_row=first_period_row, start_column=1, end_row=first_period_row + 3, end_column=1)
    sheet.merge_cells(start_row=first_period_row + 4, start_column=1, end_row=last_period_row, end_column=1)
    sheet[morning_cell] = "上午"
    sheet[afternoon_cell] = "下午"
    for cell_ref in (morning_cell, afternoon_cell):
        sheet[cell_ref].font = HEADER_FONT
        sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 7.5, "B": 9, "C": 16.5, "D": 16.5, "E": 16.5, "F": 16.5, "G": 16.5}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.print_area = f"A1:G{last_period_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    sheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0.1, footer=0.1)
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.auto_filter.ref = None


def _workbook_bytes(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_class_schedule_xlsx(state: dict[str, Any], grade: int | None = None) -> BytesIO:
    schedule = _validated_schedule(state)
    classes = [
        class_item
        for class_item in state.get("classes", [])
        if grade is None or int(class_item["grade"]) == grade
    ]
    if not classes:
        raise ValueError("所选年级没有可导出的班级课表")

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    school_name = state.get("school_name", "")
    teacher_map = {teacher["id"]: teacher["name"] for teacher in state.get("teachers", [])}
    all_slots = {f"{day['id']}-{period['id']}" for day in DAYS for period in PERIODS}

    for class_item in classes:
        lessons = schedule.get("lessons", {}).get(class_item["id"], {})
        sheet = workbook.create_sheet(_safe_sheet_title(class_item["name"], used_titles))

        def class_cell(slot: str, class_lessons: dict[str, Any] = lessons) -> str:
            lesson = class_lessons.get(slot)
            if not lesson:
                return ""
            subject = SUBJECT_BY_ID.get(lesson.get("subject_id"), {"name": lesson.get("subject_id", "")})
            teacher_name = teacher_map.get(lesson.get("teacher_id"), "")
            return str(subject["name"]) + (f"\n{teacher_name}" if teacher_name else "")

        available_slots = set(slots_for_grade(int(class_item["grade"])))
        _configure_sheet(
            sheet,
            f"{school_name}{class_item['name']}课程表",
            class_cell,
            unavailable_slots=all_slots - available_slots,
        )

    workbook.calculation.fullCalcOnLoad = True
    return _workbook_bytes(workbook)


def _scheduled_teacher_ids(state: dict[str, Any], grade: int | None = None) -> set[str]:
    schedule = _validated_schedule(state)
    class_ids = {
        item["id"]
        for item in state.get("classes", [])
        if grade is None or int(item["grade"]) == grade
    }
    return {
        lesson["teacher_id"]
        for class_id in class_ids
        for lesson in schedule.get("lessons", {}).get(class_id, {}).values()
        if lesson.get("teacher_id")
    }


def build_teacher_schedules_xlsx(
    state: dict[str, Any],
    grade: int | None = None,
    teacher_id: str | None = None,
) -> BytesIO:
    schedule = _validated_schedule(state)
    if teacher_id:
        teachers = [teacher for teacher in state.get("teachers", []) if teacher["id"] == teacher_id]
        if not teachers:
            raise ValueError("教师不存在")
    else:
        scheduled_ids = _scheduled_teacher_ids(state, grade=grade)
        teachers = [teacher for teacher in state.get("teachers", []) if teacher["id"] in scheduled_ids]
    if not teachers:
        raise ValueError("所选范围内没有已排课教师")

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    school_name = state.get("school_name", "")
    class_map = {item["id"]: item for item in state.get("classes", [])}

    for teacher in teachers:
        teacher_lessons: dict[str, tuple[str, dict[str, Any]]] = {}
        for class_id, class_lessons in schedule.get("lessons", {}).items():
            for slot, lesson in class_lessons.items():
                if lesson.get("teacher_id") == teacher["id"]:
                    teacher_lessons[slot] = (class_id, lesson)
        if not teacher_lessons:
            if teacher_id:
                raise ValueError("该教师在当前课表中没有课程")
            continue

        sheet = workbook.create_sheet(_safe_sheet_title(teacher["name"], used_titles))

        def teacher_cell(slot: str, indexed: dict[str, tuple[str, dict[str, Any]]] = teacher_lessons) -> str:
            item = indexed.get(slot)
            if not item:
                return ""
            class_id, lesson = item
            class_item = class_map.get(class_id)
            class_name = class_item["name"] if class_item else class_id
            subject = SUBJECT_BY_ID.get(lesson.get("subject_id"), {"name": lesson.get("subject_id", "")})
            return f"{class_name}\n{subject['name']}"

        _configure_sheet(
            sheet,
            f"{school_name}教师课程表",
            teacher_cell,
            subtitle=f"教师：{teacher['name']}",
        )

    if not workbook.worksheets:
        raise ValueError("所选范围内没有已排课教师")
    workbook.calculation.fullCalcOnLoad = True
    return _workbook_bytes(workbook)
